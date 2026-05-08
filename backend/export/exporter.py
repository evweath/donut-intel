"""
Export Module (F39-F42)
Exports data to XLSX, CSV, and TXT formats with configurable columns.
Logs every export to ExportRecord.
"""
import csv
import io
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from backend.config import config
from backend.database.db import session_scope
from backend.database.models import (
    Competitor,
    CompetitorProductMatch,
    ExportRecord,
    PriceHistory,
    Product,
    ProductSource,
)

logger = logging.getLogger(__name__)


def _export_dir() -> Path:
    d = Path(config.get("export", "output_dir", default="./exports")).expanduser()
    if not d.is_absolute():
        d = (Path(__file__).parent.parent.parent / d).resolve()
    d.mkdir(parents=True, exist_ok=True)
    return d


def _log_export(db_session, export_type: str, filename: str, file_path: str, scope: str, row_count: int, triggered_by: str):
    db_session.add(ExportRecord(
        export_type=export_type,
        filename=filename,
        file_path=file_path,
        scope=scope,
        row_count=row_count,
        triggered_by=triggered_by,
    ))


# --------------------------------------------------------------------------
# CSV
# --------------------------------------------------------------------------

def export_products_csv(
    product_ids: Optional[List[int]] = None,
    fields: Optional[List[str]] = None,
    triggered_by: str = "user",
) -> tuple[bytes, str]:
    """Returns (csv_bytes, filename)."""
    default_fields = ["id", "title", "manufacturer", "model_number", "sku", "price", "category", "in_stock", "sources", "created_at"]
    chosen_fields = fields or default_fields

    with session_scope() as db:
        query = db.query(Product).filter(Product.is_active == True)
        if product_ids:
            query = query.filter(Product.id.in_(product_ids))
        products = query.order_by(Product.canonical_title).all()

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=chosen_fields, extrasaction="ignore")
        writer.writeheader()

        for p in products:
            row = {
                "id": p.id,
                "title": p.canonical_title,
                "manufacturer": p.manufacturer or "",
                "model_number": p.model_number or "",
                "sku": p.sku or "",
                "price": p.price_canonical or "",
                "price_min": p.price_min or "",
                "price_max": p.price_max or "",
                "category": p.category or "",
                "subcategory": p.subcategory or "",
                "in_stock": p.in_stock,
                "sources": "|".join(s.source_site for s in p.sources if s.is_active),
                "created_at": p.created_at.isoformat() if p.created_at else "",
                "updated_at": p.updated_at.isoformat() if p.updated_at else "",
            }
            writer.writerow({f: row.get(f, "") for f in chosen_fields})

        output.seek(0)
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"donut_intel_products_{ts}.csv"
        file_path = str(_export_dir() / filename)
        data = output.getvalue().encode("utf-8")
        with open(file_path, "wb") as f:
            f.write(data)

        _log_export(db, "csv", filename, file_path, "products", len(products), triggered_by)

    return data, filename


# --------------------------------------------------------------------------
# XLSX
# --------------------------------------------------------------------------

def export_products_xlsx(
    product_ids: Optional[List[int]] = None,
    include_competitors: bool = True,
    include_price_history: bool = False,
    triggered_by: str = "user",
) -> tuple[bytes, str]:
    """Returns (xlsx_bytes, filename)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
    alt_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    with session_scope() as db:
        query = db.query(Product).filter(Product.is_active == True)
        if product_ids:
            query = query.filter(Product.id.in_(product_ids))
        products = query.order_by(Product.canonical_title).all()

        # --- Sheet 1: Master Catalog ---
        ws = wb.active
        ws.title = "Master Catalog"
        headers = ["ID", "Title", "Manufacturer", "Model #", "SKU", "Price ($)", "Category", "Subcategory", "In Stock", "Sources", "Created"]
        style_header(ws, headers)

        for row_idx, p in enumerate(products, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            values = [
                p.id, p.canonical_title, p.manufacturer or "", p.model_number or "",
                p.sku or "", p.price_canonical or 0, p.category or "", p.subcategory or "",
                "Yes" if p.in_stock else "No",
                ", ".join(s.source_site for s in p.sources if s.is_active),
                p.created_at.strftime("%Y-%m-%d") if p.created_at else "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if fill:
                    cell.fill = fill

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        # --- Sheet 2: Competitor Matches ---
        if include_competitors:
            ws2 = wb.create_sheet("Competitor Matches")
            h2 = ["Product ID", "Product Title", "Competitor", "Competitor URL", "Competitor Title",
                  "Competitor Price ($)", "Our Price ($)", "Price Diff ($)", "Match Type", "Confidence", "Scanned"]
            style_header(ws2, h2)

            matches = (
                db.query(CompetitorProductMatch)
                .filter(CompetitorProductMatch.is_active == True)
                .order_by(CompetitorProductMatch.master_product_id)
                .all()
            )

            for row_idx, m in enumerate(matches, 2):
                product = db.get(Product, m.master_product_id)
                comp = db.get(Competitor, m.competitor_id)
                our_price = product.price_canonical if product else None
                diff = round(m.competitor_price - our_price, 2) if m.competitor_price and our_price else ""
                fill = alt_fill if row_idx % 2 == 0 else None
                vals = [
                    m.master_product_id,
                    product.canonical_title if product else "",
                    comp.domain if comp else "",
                    m.competitor_url or "",
                    m.competitor_title or "",
                    m.competitor_price or 0,
                    our_price or 0,
                    diff,
                    m.match_type or "",
                    m.match_confidence or 0,
                    m.scanned_at.strftime("%Y-%m-%d") if m.scanned_at else "",
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws2.cell(row=row_idx, column=col, value=val)
                    if fill:
                        cell.fill = fill

        # --- Sheet 3: Price History ---
        if include_price_history:
            ws3 = wb.create_sheet("Price History")
            h3 = ["Match ID", "Product ID", "Competitor", "Price ($)", "In Stock", "Recorded"]
            style_header(ws3, h3)
            history = db.query(PriceHistory).order_by(PriceHistory.match_id, PriceHistory.recorded_at).all()
            for row_idx, ph in enumerate(history, 2):
                match = db.get(CompetitorProductMatch, ph.match_id)
                comp = db.get(Competitor, match.competitor_id) if match else None
                ws3.cell(row=row_idx, column=1, value=ph.match_id)
                ws3.cell(row=row_idx, column=2, value=match.master_product_id if match else "")
                ws3.cell(row=row_idx, column=3, value=comp.domain if comp else "")
                ws3.cell(row=row_idx, column=4, value=ph.price)
                ws3.cell(row=row_idx, column=5, value="Yes" if ph.in_stock else "No")
                ws3.cell(row=row_idx, column=6, value=ph.recorded_at.strftime("%Y-%m-%d %H:%M") if ph.recorded_at else "")

        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"donut_intel_{ts}.xlsx"
        file_path = str(_export_dir() / filename)

        output = io.BytesIO()
        wb.save(output)
        data = output.getvalue()

        with open(file_path, "wb") as f:
            f.write(data)

        _log_export(db, "xlsx", filename, file_path, "full_export", len(products), triggered_by)

    return data, filename


# --------------------------------------------------------------------------
# TXT
# --------------------------------------------------------------------------

def export_products_txt(
    product_ids: Optional[List[int]] = None,
    triggered_by: str = "user",
) -> tuple[bytes, str]:
    lines = []
    lines.append("=" * 80)
    lines.append("DONUT INTEL PLATFORM — PRODUCT CATALOG EXPORT")
    lines.append(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    lines.append("=" * 80)
    lines.append("")

    with session_scope() as db:
        query = db.query(Product).filter(Product.is_active == True)
        if product_ids:
            query = query.filter(Product.id.in_(product_ids))
        products = query.order_by(Product.canonical_title).all()

        for p in products:
            lines.append(f"[{p.id}] {p.canonical_title}")
            if p.manufacturer:
                lines.append(f"  Manufacturer : {p.manufacturer}")
            if p.model_number:
                lines.append(f"  Model #      : {p.model_number}")
            if p.sku:
                lines.append(f"  SKU          : {p.sku}")
            if p.price_canonical:
                lines.append(f"  Price        : ${p.price_canonical:.2f}")
            if p.category:
                lines.append(f"  Category     : {p.category}")
            lines.append(f"  In Stock     : {'Yes' if p.in_stock else 'No'}")
            active_sources = [s for s in p.sources if s.is_active]
            if active_sources:
                lines.append(f"  Sources      : {', '.join(s.source_site for s in active_sources)}")
            lines.append("")

        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"donut_intel_products_{ts}.txt"
        file_path = str(_export_dir() / filename)
        data = "\n".join(lines).encode("utf-8")
        with open(file_path, "wb") as f:
            f.write(data)

        _log_export(db, "txt", filename, file_path, "products", len(products), triggered_by)

    return data, filename


# --------------------------------------------------------------------------
# Dispatcher
# --------------------------------------------------------------------------

def export_products(fmt: str = "xlsx", triggered_by: str = "user", **kwargs):
    """Generic dispatcher used by the scheduler."""
    if fmt == "xlsx":
        return export_products_xlsx(triggered_by=triggered_by, **kwargs)
    elif fmt == "csv":
        return export_products_csv(triggered_by=triggered_by, **kwargs)
    elif fmt == "txt":
        return export_products_txt(triggered_by=triggered_by, **kwargs)
    raise ValueError(f"Unsupported export format: {fmt}")
